"""
settings.py  v1.2
==================
設定の読み書きを一元管理するモジュール。

[v1.2 変更点]
  - DEFAULT_SHIFT_HOURS に 10B（実働10h、夜勤扱い調整用）を追加
  - DEFAULT_CONSTRAINTS の月間上限時間を 188h に変更（上限以内運用に対応）
  - validate() のシフト名チェックを新体系に統一
"""

from pathlib import Path
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===========================================================================
# デフォルト設定値
# ===========================================================================

DEFAULT_ROSTER: List[str] = [
    "末吉 弘一",
    "伊藤 晶俊",
    "吉村 智",
    "南 英俊",
    "杉田 孝行",
    "山田 誠",
    "大西 信一",
    "村主 博",
    "河内 拳",
]

DEFAULT_FIXED_WORKER: str = "末吉 弘一"

DEFAULT_SHIFT_HOURS: Dict[str, int] = {
    "隊長日勤": 8,
    "日勤A":    8,
    "日勤B":    8,
    "A":        11,
    "9B":       10,
    "10B":      10,   # 時間調整用オプション夜勤（実働10h）
    "C":        11,
    "○":        0,
}

DEFAULT_CONSTRAINTS: Dict[str, int] = {
    "月間上限時間":         188,   # 上限以内運用。実態に合わせて調整
    "連勤制限":               5,
    "週休判定ウィンドウ幅":   7,
}

CONSTRAINT_DESCRIPTIONS: Dict[str, str] = {
    "月間上限時間":         "1人あたりの月間最大労働時間 (h)。この時間以内に収める。",
    "連勤制限":             "連続して勤務できる最大日数",
    "週休判定ウィンドウ幅": "週1休を判定するスライディングウィンドウの幅 (日)",
}

SHEET_NAME = "設定"


# ===========================================================================
# 設定クラス
# ===========================================================================

class Settings:
    def __init__(self):
        self.roster:       List[str]      = list(DEFAULT_ROSTER)
        self.fixed_worker: str            = DEFAULT_FIXED_WORKER
        self.shift_hours:  Dict[str, int] = dict(DEFAULT_SHIFT_HOURS)
        self.constraints:  Dict[str, int] = dict(DEFAULT_CONSTRAINTS)

    # -----------------------------------------------------------------------
    # 読み込み
    # -----------------------------------------------------------------------
    @classmethod
    def load(cls, filepath: Path) -> "Settings":
        s = cls()
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception:
            return s

        if SHEET_NAME not in wb.sheetnames:
            return s

        ws   = wb[SHEET_NAME]
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]

        s._parse_roster(rows)
        s._parse_fixed_worker(rows)
        s._parse_shift_hours(rows)
        s._parse_constraints(rows)
        return s

    def _find_block(self, rows: list, header: str) -> Optional[int]:
        for i, row in enumerate(rows):
            if row and str(row[0]).strip() == header:
                return i
        return None

    def _parse_roster(self, rows: list):
        idx = self._find_block(rows, "■ 従業員名簿")
        if idx is None:
            return
        roster = []
        for row in rows[idx + 2:]:
            if not row or row[0] is None or str(row[0]).strip() == "":
                break
            name = str(row[0]).strip()
            if name:
                roster.append(name)
        if roster:
            self.roster = roster

    def _parse_fixed_worker(self, rows: list):
        idx = self._find_block(rows, "■ 固定ワーカー設定")
        if idx is None:
            return
        for row in rows[idx + 3:]:
            if not row or row[0] is None or str(row[0]).strip() == "":
                break
            val = row[1] if len(row) > 1 else None
            if val is not None and str(val).strip():
                self.fixed_worker = str(val).strip()
            break

    def _parse_shift_hours(self, rows: list):
        idx = self._find_block(rows, "■ シフト種類・勤務時間")
        if idx is None:
            return
        shift_hours = {}
        for row in rows[idx + 2:]:
            if not row or row[0] is None or str(row[0]).strip() == "":
                break
            shift = str(row[0]).strip()
            try:
                hours = int(row[1]) if len(row) > 1 and row[1] is not None else 0
                shift_hours[shift] = hours
            except (ValueError, TypeError):
                pass
        if shift_hours:
            self.shift_hours = shift_hours

    def _parse_constraints(self, rows: list):
        idx = self._find_block(rows, "■ 制約パラメータ")
        if idx is None:
            return
        for row in rows[idx + 2:]:
            if not row or row[0] is None or str(row[0]).strip() == "":
                break
            key = str(row[0]).strip()
            try:
                val = int(row[1]) if len(row) > 1 and row[1] is not None else None
                if key in self.constraints and val is not None:
                    self.constraints[key] = val
            except (ValueError, TypeError):
                pass

    # -----------------------------------------------------------------------
    # 保存
    # -----------------------------------------------------------------------
    def save(self, filepath: Path):
        try:
            wb = openpyxl.load_workbook(filepath)
        except Exception:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME, 0)

        writer = _SheetWriter(ws)
        writer.write_roster(self.roster)
        writer.write_fixed_worker(self.fixed_worker)
        writer.write_shift_hours(self.shift_hours)
        writer.write_constraints(self.constraints)
        writer.adjust_columns()
        wb.save(filepath)

    # -----------------------------------------------------------------------
    # バリデーション
    # -----------------------------------------------------------------------
    def validate(self) -> List[str]:
        errors = []

        if not self.roster:
            errors.append("従業員名簿が空です。")

        if self.fixed_worker and self.fixed_worker not in self.roster:
            errors.append(
                f"固定ワーカー「{self.fixed_worker}」が従業員名簿に存在しません。"
            )

        # optimizer が毎日必須とするシフトの存在確認
        required = ["隊長日勤", "日勤A", "日勤B", "A", "9B", "C", "○"]
        for shift in required:
            if shift not in self.shift_hours:
                errors.append(f"必須シフト「{shift}」がシフト種類に定義されていません。")

        # 1日5シフト必要（日勤A/B + A/9B/C）、固定ワーカー除いて5名以上
        daily_required = 5
        n_shift_workers = len([w for w in self.roster if w != self.fixed_worker])
        if n_shift_workers < daily_required:
            errors.append(
                f"シフト対象者({n_shift_workers}名) < 1日の必要人数({daily_required}名)。"
                f"従業員を追加してください。"
            )

        return errors

    # -----------------------------------------------------------------------
    # プロパティ
    # -----------------------------------------------------------------------
    @property
    def night_shifts(self) -> List[str]:
        """夜勤扱いシフト（10B含む）"""
        return ["A", "9B", "10B", "C"]

    @property
    def shift_types(self) -> List[str]:
        non_rest = [s for s in self.shift_hours if s != "○"]
        return non_rest + ["○"]


# ===========================================================================
# Excel書き込みヘルパー
# ===========================================================================

class _SheetWriter:
    COLOR_HEADER_BG = "1A1A2E"
    COLOR_HEADER_FG = "FFFFFF"
    COLOR_COL_BG    = "3A3A5C"
    COLOR_COL_FG    = "FFFFFF"
    COLOR_DATA_ALT  = "F0F4FF"
    COLOR_NOTE      = "888888"

    def __init__(self, ws):
        self.ws  = ws
        self.row = 1

    def write_roster(self, roster: List[str]):
        self._section_header("■ 従業員名簿")
        self._col_headers(["名前"])
        for i, name in enumerate(roster):
            self._data_row([name], i)
        self.row += 1

    def write_fixed_worker(self, fixed_worker: str):
        self._section_header("■ 固定ワーカー設定")
        self._note_row("平日は隊長日勤固定・土日は○固定となる従業員を指定（1名のみ・空欄で無効）")
        self._col_headers(["設定項目", "値"])
        self._data_row(["固定ワーカー名", fixed_worker], 0)
        self.row += 1

    def write_shift_hours(self, shift_hours: Dict[str, int]):
        self._section_header("■ シフト種類・勤務時間")
        self._note_row("10B は毎日必須ではなく月間時間調整用オプション夜勤（実働10h）")
        self._col_headers(["シフト名", "勤務時間 (h)"])
        for i, (shift, hours) in enumerate(shift_hours.items()):
            self._data_row([shift, hours], i)
        self.row += 1

    def write_constraints(self, constraints: Dict[str, int]):
        self._section_header("■ 制約パラメータ")
        self._col_headers(["パラメータ名", "値", "説明"])
        for i, (key, val) in enumerate(constraints.items()):
            desc = CONSTRAINT_DESCRIPTIONS.get(key, "")
            self._data_row([key, val, desc], i)
        self.row += 1

    def _section_header(self, title: str):
        cell = self.ws.cell(self.row, 1, title)
        cell.font      = Font(bold=True, color=self.COLOR_HEADER_FG, size=11)
        cell.fill      = PatternFill("solid", fgColor=self.COLOR_HEADER_BG)
        cell.alignment = Alignment(vertical="center", indent=1)
        for c in range(2, 4):
            self.ws.cell(self.row, c).fill = PatternFill("solid", fgColor=self.COLOR_HEADER_BG)
        self.ws.row_dimensions[self.row].height = 22
        self.row += 1

    def _note_row(self, text: str):
        cell = self.ws.cell(self.row, 1, f"  ※ {text}")
        cell.font      = Font(color=self.COLOR_NOTE, italic=True, size=9)
        cell.alignment = Alignment(vertical="center")
        self.row += 1

    def _col_headers(self, headers: List[str]):
        for c, h in enumerate(headers, start=1):
            cell = self.ws.cell(self.row, c, h)
            cell.font      = Font(bold=True, color=self.COLOR_COL_FG)
            cell.fill      = PatternFill("solid", fgColor=self.COLOR_COL_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        self.ws.row_dimensions[self.row].height = 18
        self.row += 1

    def _data_row(self, values: list, index: int):
        bg     = self.COLOR_DATA_ALT if index % 2 == 1 else None
        border = Border(bottom=Side(style="hair", color="DDDDDD"))
        for c, val in enumerate(values, start=1):
            cell = self.ws.cell(self.row, c, val)
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
            cell.border    = border
            cell.alignment = Alignment(vertical="center", indent=1)
        self.row += 1

    def adjust_columns(self):
        col_widths = {}
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                col    = cell.column
                length = len(str(cell.value)) * 2
                col_widths[col] = max(col_widths.get(col, 10), min(length, 50))
        for col, width in col_widths.items():
            self.ws.column_dimensions[get_column_letter(col)].width = width + 2
